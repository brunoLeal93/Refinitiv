# It make file to upload on Customer Zone
# to client can see the new listed RICs and delisted RICs  


from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import datetime
from CreateNew import newCaz
from Config import createDir

def contRowFill(wb):
    i = 3
    print(wb['INPUT SHEET'].cell(row=2, column=1).value)
    while wb['INPUT SHEET'].cell(row=i, column=1).value is not None:
        i += 1
    return i


def createCAZ(df, tp, date):
    localSaveCaz = createDir(['DATA', date, 'CAZ'])
    EffectiveDate = date
    try:
        wb = load_workbook(localSaveCaz+'/New RICs CAZ - Venezuela '+EffectiveDate+'.xlsx')
    except:
        wb = newCaz()
    
    i = contRowFill(wb)
    # print(i)d
    print(df)
    if tp == 'Bonds':
        for index, row in df.iterrows():
            wb['INPUT SHEET'].cell(row=i, column=1).value = 'Create'
            wb['INPUT SHEET'].cell(row=i, column=3).value = 'Add'
            wb['INPUT SHEET'].cell(row=i, column=4).value = EffectiveDate
            wb['INPUT SHEET'].cell(row=i, column=6).value = row['SYMBOL']
            wb['INPUT SHEET'].cell(row=i, column=8).value = row['RIC']
            wb['INPUT SHEET'].cell(row=i, column=10).value = row['OFFCL_CODE']
            wb['INPUT SHEET'].cell(row=i, column=11).value = 'Ticker'
            wb['INPUT SHEET'].cell(row=i, column=13).value = row['SYMBOL']
            wb['INPUT SHEET'].cell(row=i, column=17).value = 'CCS'
            wb['INPUT SHEET'].cell(row=i, column=18).value = 'BON'
            i=i+1
        wb.save(localSaveCaz+'/New RICs CAZ - Venezuela '+EffectiveDate+'.xlsx')
        wb.close()
        return localSaveCaz+'/New RICs CAZ - Venezuela '+EffectiveDate+'.xlsx'
        

    if tp == 'Equities':
        for index, row in df.iterrows():
            wb['INPUT SHEET'].cell(row=i, column=1).value = 'Create'
            wb['INPUT SHEET'].cell(row=i, column=3).value = 'Add'
            wb['INPUT SHEET'].cell(row=i, column=4).value = EffectiveDate
            wb['INPUT SHEET'].cell(row=i, column=6).value = row['DSPLY_NAME']
            wb['INPUT SHEET'].cell(row=i, column=8).value = row['RIC']
            wb['INPUT SHEET'].cell(row=i, column=10).value = row['OFFCL_CODE']
            wb['INPUT SHEET'].cell(row=i, column=11).value = 'Ticker'
            wb['INPUT SHEET'].cell(row=i, column=13).value = row['SYMBOL']
            wb['INPUT SHEET'].cell(row=i, column=17).value = 'CCS'
            wb['INPUT SHEET'].cell(row=i, column=18).value = 'EQI'
            i=i+1
        wb.save(localSaveCaz+'/New RICs CAZ - Venezuela '+EffectiveDate+'.xlsx')
        wb.close()
        return localSaveCaz+'/New RICs CAZ - Venezuela '+EffectiveDate+'.xlsx'
    

