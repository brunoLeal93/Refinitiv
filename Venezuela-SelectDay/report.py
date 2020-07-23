from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import datetime
import os
from Config import createDir
from CreateNew import newReport

def contRowFill(wb):
    i = 6
    while wb['Sheet1'].cell(row=i, column=2).value is not None:
        i += 1
    return i


def createReport(df, tp, date):
    localSaveReport = createDir(['DATA', date, 'Reports'])
    EffectiveDate = date
    try:
        wb = load_workbook(localSaveReport+'/New RICs Report - Venezuela '+EffectiveDate+'.xlsx')
    except:
        wb = newReport()
    
    i = contRowFill(wb)

    if tp == 'Bonds':
        for index, row in df.iterrows():
            wb['Sheet1'].cell(row=i, column=2).value = row['SYMBOL']
            wb['Sheet1'].cell(row=i, column=3).value = row['DSPLY_NAME']
            wb['Sheet1'].cell(row=i, column=4).value = row['RIC']
            wb['Sheet1'].cell(row=i, column=5).value = row['ORG_NAME']
            wb['Sheet1'].cell(row=i, column=6).value = 'ADD'
            wb['Sheet1'].cell(row=i, column=7).value = row['OFFCL_CODE']
            wb['Sheet1'].cell(row=i, column=8).value = EffectiveDate
            i=i+1
        wb.save(localSaveReport+'/New RICs Report - Venezuela '+EffectiveDate+'.xlsx')
        wb.close()
        return localSaveReport+'/New RICs Report - Venezuela '+EffectiveDate+'.xlsx'

    elif tp == 'Equities':
        for index, row in df.iterrows():
            wb['Sheet1'].cell(row=i, column=2).value = row['SYMBOL']
            wb['Sheet1'].cell(row=i, column=3).value = row['DSPLY_NAME']
            wb['Sheet1'].cell(row=i, column=4).value = row['RIC']
            wb['Sheet1'].cell(row=i, column=5).value = row['ORG_NAME']
            wb['Sheet1'].cell(row=i, column=6).value = 'ADD'
            wb['Sheet1'].cell(row=i, column=7).value = row['OFFCL_CODE']
            wb['Sheet1'].cell(row=i, column=8).value = EffectiveDate
            i=i+1
        wb.save(localSaveReport+'/New RICs Report - Venezuela '+EffectiveDate+'.xlsx')
        wb.close()
        return localSaveReport+'/New RICs Report - Venezuela '+EffectiveDate+'.xlsx'
    

