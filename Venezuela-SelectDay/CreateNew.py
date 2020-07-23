from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.styles.colors import RED, BLUE, BLACK

def newReport():
    title = NamedStyle(name='title')
    title.font = Font(name='Ariel',
                    size=10,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color=RED)
    title.alignment = Alignment(horizontal='left',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    exch = NamedStyle(name='exch')
    exch.font = Font(name='Ariel',
                    size=10,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color=BLUE)
    exch.alignment = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    head = NamedStyle(name='head')
    head.font = Font(name='Ariel',
                    size=8,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color=BLACK)
    head.alignment = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)
    head.fill = PatternFill(fill_type='solid',
                start_color='ccffcc')

    head.border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    wb = Workbook()

    ws1 = wb.create_sheet("Sheet1", 0)


    ws1['B1'] = 'ADDITIONS ON CCSE'
    ws1['B3'] = 'CCSE'
    ws1['B5'] = 'TICKER'
    ws1['C5'] = 'DISPLAY NAME'
    ws1['D5'] = 'RIC'
    ws1['E5'] = 'FULL NAME'
    ws1['F5'] = 'STATUS'
    ws1['G5'] = 'ISIN'
    ws1['H5'] = 'ISSUE DATE'

    ws1['B1'].style = title
    ws1['B3'].style = exch
    ws1['B5'].style = head
    ws1['C5'].style = head
    ws1['D5'].style = head
    ws1['E5'].style = head
    ws1['F5'].style = head
    ws1['G5'].style = head
    ws1['H5'].style = head

    return wb

def newCaz():
    wb = Workbook()

    ws1 = wb.create_sheet("INPUT SHEET", 0)

    ws1['A2'] = 'EventAction'
    ws1['B2'] = 'RicSeqId'
    ws1['C2'] = 'Change Type'
    ws1['D2'] = 'Date'
    ws1['E2'] = 'Description Was'
    ws1['F2'] = 'Description Now'
    ws1['G2'] = 'RICwas'
    ws1['H2'] = 'RICnow'
    ws1['I2'] = 'ISINwas'
    ws1['J2'] = 'ISINnow'
    ws1['K2'] = '2ndID'
    ws1['L2'] = '2ndwas'
    ws1['M2'] = '2ndnow'
    ws1['N2'] = 'ThomsonWas'
    ws1['O2'] = 'ThomsonNow'
    ws1['P1'] = ''
    ws1['Q2'] = 'Exchange'
    ws1['R2'] = 'Asset'

    return wb